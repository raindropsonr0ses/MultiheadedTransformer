#DataPreprocessingSegmentHumanoid

import re
import numpy as np
import torch
from torch.utils.data import Dataset, DataLoader
from openpyxl import load_workbook

class fNIRSDataset(Dataset):
    """
    Custom dataset for fNIRS HRF data from an Excel workbook.

    Each sample corresponds to one subject–event combination (one column from the headers)
    and returns (fNIRS_data, event_label).

    fNIRS_data is a tensor of shape [2, num_channels, target_length]:
      - 2: two signal types (0: HbO; 1: HbR)
      - num_channels: expected 24
      - target_length: number of time points after sliding-window processing

    Event label mapping:
      - If header has {"S", "F", "H"} (tertiary): {"S": 0, "F": 1, "H": 2}
      - Otherwise (binary, only {"F","H"}): {"F": 0, "H": 1}

    Sliding-window processing:
      Tertiary (has S):
         - S: already 4549 points;
         - F: valid length 4801 → windows [0:4549] and [252:252+4549] averaged → 4549 points;
         - H: valid length 6801 → windows [0:4549] and [2252:2252+4549] averaged.
      Binary (only F and H):
         - F: valid length 4801 → take first 4801 points;
         - H: valid length 6801 → windows [0:4801] and [2000:2000+4801] averaged.
    """
    def __init__(self, excel_file, split="train"):
        self.excel_file = excel_file
        wb = load_workbook(excel_file, read_only=True)
        all_sheet_names = wb.sheetnames

        # Group sheets by channel id using regex to extract "number,number"
        channel_dict = {}
        for sheet_name in all_sheet_names:
            name_lower = sheet_name.lower()
            measurement = None
            if "hbo" in name_lower:
                measurement = "HbO"
            elif "hbr" in name_lower:
                measurement = "HbR"
            else:
                continue
            match = re.search(r'(\d+,\d+)', sheet_name)
            if not match:
                continue
            channel_id = match.group(1)
            channel_dict.setdefault(channel_id, {})[measurement] = sheet_name

        valid_channels = {cid: sheets for cid, sheets in channel_dict.items()
                          if "HbO" in sheets and "HbR" in sheets}
        self.channels = sorted(valid_channels.keys())
        self.num_channels = len(self.channels)
        if self.num_channels == 0:
            raise ValueError("No valid channels found. Ensure sheet names contain 'HbO' or 'HbR' and a channel id in the format 'number,number'.")
        self.hb_sheets = [(valid_channels[cid]["HbO"], valid_channels[cid]["HbR"]) for cid in self.channels]

        # Extract header information from one HbO sheet.
        sample_sheet = wb[self.hb_sheets[0][0]]
        rows = list(sample_sheet.iter_rows(values_only=True))
        header1 = list(rows[0])  # event labels
        header2 = list(rows[1])  # subject identifiers (not used for classification)
        expected_events = {"S", "F", "H"}
        # If first cell not an expected event, assume timestamp column.
        if header1[0] not in expected_events:
            header1 = header1[1:]
            header2 = header2[1:]
            self.drop_first = True
        else:
            self.drop_first = False
        self.sample_headers = list(zip(header1, header2))

        # Split subjects into train (16 subjects) and test (4 subjects).
        all_subjects = sorted(list(set([subj for (_, subj) in self.sample_headers])))
        if len(all_subjects) != 20:
            print(f"Warning: Expected 20 subjects but found {len(all_subjects)}.")
        if split == "train":
            selected_subjects = all_subjects[:16]
        elif split == "test":
            selected_subjects = all_subjects[16:]
        elif split in ["all", "validation"]:
            selected_subjects = all_subjects
        else:
            raise ValueError("Invalid split type. Must be 'train' or 'test'.")
        self.samples_meta = []
        for col_idx, (event, subj) in enumerate(self.sample_headers):
            if subj in selected_subjects:
                self.samples_meta.append((event, subj, col_idx))
        self.sampling_points = len(rows) - 2  # maximum available time points in sheet
        wb.close()

        # Set event mapping and target length based on header content.
        if expected_events == {"S", "F", "H"}:
            self.event_map = {"S": 0, "F": 1, "H": 2}
            self.target_length = 4549
            self.offset_F = 252      # 4801 - 4549
            self.offset_H = 2252     # 6801 - 4549
        else:  # binary: only F and H
            self.event_map = {"F": 0, "H": 1}
            self.target_length = 4801
            self.offset_H = 2000     # 6801 - 4801

    def __len__(self):
        return len(self.samples_meta)

    def __getitem__(self, idx):
        event, subj, col_idx = self.samples_meta[idx]
        actual_col_idx = col_idx + 1 if self.drop_first else col_idx

        wb = load_workbook(self.excel_file, read_only=True)
        hbO_list = []
        hbR_list = []
        for hbO_sheet_name, hbR_sheet_name in self.hb_sheets:
            # Extract HbO data.
            sheet_hbO = wb[hbO_sheet_name]
            col_vals_hbO = []
            for row in sheet_hbO.iter_rows(min_row=3, values_only=True):
                val = row[actual_col_idx] if actual_col_idx < len(row) else 0
                if val is None:
                    val = 0
                col_vals_hbO.append(val * 1_000_000)
            hbO_list.append(np.array(col_vals_hbO, dtype=np.float32))
            # Extract HbR data.
            sheet_hbR = wb[hbR_sheet_name]
            col_vals_hbR = []
            for row in sheet_hbR.iter_rows(min_row=3, values_only=True):
                val = row[actual_col_idx] if actual_col_idx < len(row) else 0
                if val is None:
                    val = 0
                col_vals_hbR.append(val * 1_000_000)
            hbR_list.append(np.array(col_vals_hbR, dtype=np.float32))
        wb.close()

        HbO_data = np.stack(hbO_list, axis=0)  # [num_channels, T]
        HbR_data = np.stack(hbR_list, axis=0)  # [num_channels, T]
        fNIRS_data = np.stack([HbO_data, HbR_data], axis=0)  # [2, num_channels, T]
        fNIRS_data = torch.tensor(fNIRS_data, dtype=torch.float32)

        # Sliding-window processing based on whether header has "S" (tertiary) or not (binary).
        if "S" in self.event_map:  # tertiary: target_length = 4549
            if event == "S":
                new_data = fNIRS_data[:, :, :self.target_length]
            elif event == "F":
                window1 = fNIRS_data[:, :, :self.target_length]
                window2 = fNIRS_data[:, :, self.offset_F:self.offset_F+self.target_length]
                new_data = (window1 + window2) / 2.0
            elif event == "H":
                window1 = fNIRS_data[:, :, :self.target_length]
                window2 = fNIRS_data[:, :, self.offset_H:self.offset_H+self.target_length]
                new_data = (window1 + window2) / 2.0
            else:
                new_data = fNIRS_data[:, :, :self.target_length]
        else:  # binary: target_length = 4801
            if event == "F":
                new_data = fNIRS_data[:, :, :self.target_length]
            elif event == "H":
                window1 = fNIRS_data[:, :, :self.target_length]
                window2 = fNIRS_data[:, :, self.offset_H:self.offset_H+self.target_length]
                new_data = (window1 + window2) / 2.0
            else:
                new_data = fNIRS_data[:, :, :self.target_length]

        event_label = self.event_map.get(event, -1)
        return new_data, event_label

excel_path = '/content/Subjectwise Conc (GLM+no MA) with S.xlsx'
validation_excel_path = '/content/Subjectwise Validation Conc (GLM+no MA) with S.xlsx'
train_dataset = fNIRSDataset(excel_path, split="train")
val_dataset   = fNIRSDataset(validation_excel_path, split="all")
test_dataset  = fNIRSDataset(excel_path, split="test")

print("Train samples:", len(train_dataset))
print("Validation samples:", len(val_dataset))
print("Test samples:", len(test_dataset))

sample_data, event_lbl = train_dataset[0]
print("fNIRS data shape:", sample_data.shape)
print("Labels -> Event: {}".format(event_lbl))
